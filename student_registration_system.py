from tkinter import *
from datetime import date
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)

radio = IntVar(value=0)

# Check if Excel file exists
file = pathlib.Path('Student_data.xlsx')
if not file.exists():
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father's Name"
    sheet['J1'] = "Mother's Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"
    file.save('Student_data.xlsx')

#Registration No.
def registration_no():
    pass

#Exit Window
def Exit():
    root.destroy()

#Search for record
def search_record():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == Search.get():
            # Populate fields with the record
            Registration.set(sheet.cell(row=row, column=1).value)
            Name.set(sheet.cell(row=row, column=2).value)
            # Continue for other fields
            break
    else:
        messagebox.showinfo("Search", "No record found.")


#Shows image
def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Select image file",
        filetypes=(("JPG File", "*.jpg"), ("PNG File", "*.png"), ("All files", "*.*"))
    )

    if filename:  # Ensure a file is selected
        try:
            img = Image.open(filename)
            resized_image = img.resize((190, 190))
            photo2 = ImageTk.PhotoImage(resized_image)
            lbl.config(image=photo2)
            lbl.image = photo2
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open image: {e}")
    else:
        messagebox.showinfo("No File", "No image file was selected.")

#Save record
def save_record():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    sheet.append([Registration.get(), Name.get(), Class.get(), "Male" if radio.get() == 1 else "Female",
                  DOB.get(), Date.get(), Religion.get(), Skill.get(),
                  F_Name.get(), M_Name.get(), Father_Occupation.get(), Mother_Occupation.get()])
    file.save('Student_data.xlsx')
    messagebox.showinfo("Success", "Record Saved Successfully!")


#Registration No.
def registration_no():
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value

    try:
        Registration.set(max_row_value+1)

    except:
        Registration.set("1")

#Clear
def Clear():
    global img
    # Reset all input variables
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set("Select Class")
    
    # Re-generate registration number if applicable
    registration_no()
    
    # Re-enable the save button
    saveButton.config(state='normal')
    
    # Reset the image
    img1=PhotoImage(file='C:/Users/realm/OneDrive/Desktop/Login System with mySQL/Images/upload profile photo.png')
    lbl.config(image=img1)
    lbl.image=img1

    img=""

def search():
    text = Search.get()

    Clear()
    saveButton.config(state='disable')

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name=row[0]
            reg_no_position=(str(name))[14:-1]
            reg_number=str(name)[15:-1]

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid registration number!!!")

    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value
            
    

#Save
def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()

    # Ensure gender is selected
    if radio.get() == 0:  # No radio button selected
        messagebox.showerror("Error", "Select Gender!")
        return  # Stop execution of the function

    G1 = "Male" if radio.get() == 1 else "Female"

    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()

    if N1=="" or C1=="Select Class" or D2=="" or Re1=="" or S1=="" or fathername=="" or mothername=="" or F1=="" or M1=="":
        messagebox.showerror("error","Few Data is missing!")
    else:
        file=openpyxl.load_workbook('Student_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=C1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=D2)
        sheet.cell(column=6,row=sheet.max_row,value=D1)
        sheet.cell(column=7,row=sheet.max_row,value=Re1)
        sheet.cell(column=8,row=sheet.max_row,value=S1)
        sheet.cell(column=9,row=sheet.max_row,value=fathername)
        sheet.cell(column=10,row=sheet.max_row,value=mothername)
        sheet.cell(column=11,row=sheet.max_row,value=F1)
        sheet.cell(column=12,row=sheet.max_row,value=M1)
        file.save(r'Student_data.xlsx')

        try:
            img.save("C:/Users/realm/OneDrive/Desktop/Login System with mySQL/student images/" + str(R1) + ".jpg")
        except AttributeError:
            # This exception is raised if `img` is None or does not have a `save` method
            message.showinfo("Info", "Profile picture is not available!")
        except Exception as e:
            # Catch other exceptions to ensure the program doesn't crash
            message.showerror("Error", f"An error occurred while saving the image: {e}")


        messagebox.showinfo("info","Data successfully entered!")

        Clear()

        registration_no()


#Gender
def selection():
    global gender
    value=radio.get()
    if value == 1:
        gender="Male"
        print(gender)
    else:
        gender="Female"
        print(gender)

# Top Frames with Contact and Title
Label(root, text="Email: zwelakhetechnologies@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg="#c36464", fg='#fff', font='arial 20 bold').pack(side=TOP, fill=X, padx=(20,0))

# Search box and button
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font="arial 20").place(x=820, y=70)

imageicon3 = PhotoImage(file="C:/Users/realm/OneDrive/Desktop/Login System with mySQL/Images/search.png").subsample(20, 35)
Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, width=123, bg="#68ddfa", font="arial 13 bold", command=search)
Srch.place(x=1060, y=70)

# Update button
imageicon4 = PhotoImage(file="C:/Users/realm/OneDrive/Desktop/Login System with mySQL/Images/Layer 4.png").subsample(8, 10)
update_button = Button(root, image=imageicon4, bg="#c36464")
update_button.place(x=110, y=64)

# Registration Number and Date
Label(root, text="Registration No:", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Date:", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)

# Auto-fill today's date
today = date.today()
d1 = today.strftime("%d%m%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)
Date.set(d1)

# Student Details Section
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Full Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Date of Birth:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Gender:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj, text="Class:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Religion:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Skills:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160, y=50)

DOB = StringVar()
dob_entry = Entry(obj, textvariable=DOB, width=20, font="arial 10")
dob_entry.place(x=160, y=100)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg)
R1.place(x=150, y=150)
R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg)
R2.place(x=200, y=150)

Religion = StringVar()
religion_entry = Entry(obj, textvariable=Religion, width=20, font="arial 10")
religion_entry.place(x=630, y=100)

Skill = StringVar()
skill_entry = Entry(obj, textvariable=Skill, width=20, font="arial 10")
skill_entry.place(x=630, y=150)

Class = Combobox(obj, values=[str(i) for i in range(1, 13)], font="Roboto 10", width=17, state="r")
Class.place(x=630, y=50)
Class.set("Select Class")

# Parent Details Section
obj2 = LabelFrame(root, text="Parent's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Father's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)

F_Name = StringVar()
f_entry = Entry(obj2, textvariable=F_Name, width=20, font="arial 10")
f_entry.place(x=160, y=50)

Father_Occupation = StringVar()
FO_entry = Entry(obj2, textvariable=Father_Occupation, width=20, font="arial 10")
FO_entry.place(x=160, y=100)

Label(obj2, text="Mother's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)

Mother_Name = StringVar()
M_entry = Entry(obj2, textvariable=Mother_Name, width=20, font="arial 10")
M_entry.place(x=630, y=50)

Mother_Occupation = StringVar()
MO_entry = Entry(obj2, textvariable=Mother_Occupation, width=20, font="arial 10")
MO_entry.place(x=630, y=100)

# Image Frame
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

default_img = PhotoImage(file="C:/Users/realm/OneDrive/Desktop/Login System with mySQL/Images/upload profile photo.png").subsample(4, 4)
img = default_img
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

# Buttons
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue",command=showimage).place(x=1000, y=370)
Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen",command=Save).place(x=1000, y=450)
Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink",command=Clear).place(x=1000, y=530)
Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey").place(x=1000, y=610)

root.mainloop()


